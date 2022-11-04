from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QHeaderView, QVBoxLayout, QWidget, QTableWidget
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os
import matplotlib.patches as mpatches

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(821, 562)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.frame_3 = QtWidgets.QFrame(self.centralwidget)
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.frame_3)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.frame = QtWidgets.QFrame(self.frame_3)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout(self.frame)
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.maintableURL = QtWidgets.QLabel(self.frame)
        self.maintableURL.setObjectName("maintableURL")
        self.horizontalLayout_6.addWidget(self.maintableURL)
        self.verticalLayout_5.addWidget(self.frame)
        self.tabWidget = QtWidgets.QTabWidget(self.frame_3)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout(self.tab)
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.maintable = QtWidgets.QTableWidget(self.tab)
        self.maintable.setObjectName("maintable")
        self.maintable.setColumnCount(1)
        self.maintable.setRowCount(1)
        item = QtWidgets.QTableWidgetItem()
        self.maintable.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.maintable.setHorizontalHeaderItem(0, item)
        self.verticalLayout_6.addWidget(self.maintable)
        self.tabWidget.addTab(self.tab, "")
        self.verticalLayout_5.addWidget(self.tabWidget)
        self.frame_7 = QtWidgets.QFrame(self.frame_3)
        self.frame_7.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_7.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_7.setObjectName("frame_7")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.frame_7)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.delayURL = QtWidgets.QLabel(self.frame_7)
        self.delayURL.setObjectName("delayURL")
        self.verticalLayout_3.addWidget(self.delayURL)
        self.verticalLayout_5.addWidget(self.frame_7)
        self.frame_14 = QtWidgets.QFrame(self.frame_3)
        self.frame_14.setMaximumSize(QtCore.QSize(16777215, 70))
        self.frame_14.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_14.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_14.setObjectName("frame_14")
        self.verticalLayout_10 = QtWidgets.QVBoxLayout(self.frame_14)
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.delaytable = QtWidgets.QTableWidget(self.frame_14)
        self.delaytable.setMaximumSize(QtCore.QSize(16777215, 50))
        self.delaytable.setObjectName("delaytable")
        self.delaytable.setColumnCount(1)
        self.delaytable.setRowCount(1)
        item = QtWidgets.QTableWidgetItem()
        self.delaytable.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.delaytable.setHorizontalHeaderItem(0, item)
        self.verticalLayout_10.addWidget(self.delaytable)
        self.verticalLayout_5.addWidget(self.frame_14)
        self.horizontalLayout.addWidget(self.frame_3)
        self.frame_2 = QtWidgets.QFrame(self.centralwidget)
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.frame_2)
        self.verticalLayout.setObjectName("verticalLayout")
        self.frame_5 = QtWidgets.QFrame(self.frame_2)
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        self.verticalLayout_7 = QtWidgets.QVBoxLayout(self.frame_5)
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.Cmax = QtWidgets.QLabel(self.frame_5)
        font = QtGui.QFont()
        font.setPointSize(16)
        self.Cmax.setFont(font)
        self.Cmax.setObjectName("Cmax")
        self.verticalLayout_7.addWidget(self.Cmax)
        self.verticalLayout.addWidget(self.frame_5)
        self.frame_9 = QtWidgets.QFrame(self.frame_2)
        self.frame_9.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_9.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_9.setObjectName("frame_9")
        self.verticalLayout_12 = QtWidgets.QVBoxLayout(self.frame_9)
        self.verticalLayout_12.setObjectName("verticalLayout_12")
        self.frame_4 = QtWidgets.QFrame(self.frame_9)
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.frame_4)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.prep = QtWidgets.QCheckBox(self.frame_4)
        self.prep.setObjectName("prep")
        self.verticalLayout_2.addWidget(self.prep)
        self.delay = QtWidgets.QCheckBox(self.frame_4)
        self.delay.setObjectName("delay")
        self.verticalLayout_2.addWidget(self.delay)
        self.verticalLayout_12.addWidget(self.frame_4)
        self.frame_17 = QtWidgets.QFrame(self.frame_9)
        self.frame_17.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_17.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_17.setObjectName("frame_17")
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout(self.frame_17)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.frame_11 = QtWidgets.QFrame(self.frame_17)
        self.frame_11.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_11.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_11.setObjectName("frame_11")
        self.verticalLayout_9 = QtWidgets.QVBoxLayout(self.frame_11)
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.label = QtWidgets.QLabel(self.frame_11)
        self.label.setObjectName("label")
        self.verticalLayout_9.addWidget(self.label)
        self.label_2 = QtWidgets.QLabel(self.frame_11)
        self.label_2.setObjectName("label_2")
        self.verticalLayout_9.addWidget(self.label_2)
        self.horizontalLayout_5.addWidget(self.frame_11)
        self.frame_6 = QtWidgets.QFrame(self.frame_17)
        self.frame_6.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_6.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_6.setObjectName("frame_6")
        self.verticalLayout_8 = QtWidgets.QVBoxLayout(self.frame_6)
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.spinBox = QtWidgets.QSpinBox(self.frame_6)
        self.spinBox.setObjectName("spinBox")
        self.verticalLayout_8.addWidget(self.spinBox)
        self.spinBox_2 = QtWidgets.QSpinBox(self.frame_6)
        self.spinBox_2.setObjectName("spinBox_2")
        self.verticalLayout_8.addWidget(self.spinBox_2)
        self.horizontalLayout_5.addWidget(self.frame_6)
        self.verticalLayout_12.addWidget(self.frame_17)
        self.frame_15 = QtWidgets.QFrame(self.frame_9)
        self.frame_15.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_15.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_15.setObjectName("frame_15")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout(self.frame_15)
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.excelmaintableButton = QtWidgets.QPushButton(self.frame_15)
        self.excelmaintableButton.setObjectName("excelmaintableButton")
        self.verticalLayout_11.addWidget(self.excelmaintableButton)
        self.exceldelaytableButton = QtWidgets.QPushButton(self.frame_15)
        self.exceldelaytableButton.setObjectName("exceldelaytableButton")
        self.verticalLayout_11.addWidget(self.exceldelaytableButton)
        self.verticalLayout_12.addWidget(self.frame_15)
        self.frame_18 = QtWidgets.QFrame(self.frame_9)
        self.frame_18.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_18.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_18.setObjectName("frame_18")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.frame_18)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.FillButton = QtWidgets.QPushButton(self.frame_18)
        self.FillButton.setObjectName("FillButton")
        self.horizontalLayout_2.addWidget(self.FillButton)
        self.verticalLayout_12.addWidget(self.frame_18)
        self.verticalLayout.addWidget(self.frame_9)
        self.frame_10 = QtWidgets.QFrame(self.frame_2)
        self.frame_10.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_10.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_10.setObjectName("frame_10")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout(self.frame_10)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.frame_16 = QtWidgets.QFrame(self.frame_10)
        self.frame_16.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_16.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_16.setObjectName("frame_16")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(self.frame_16)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.comboBox = QtWidgets.QComboBox(self.frame_16)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.horizontalLayout_4.addWidget(self.comboBox)
        self.GoButton = QtWidgets.QPushButton(self.frame_16)
        self.GoButton.setObjectName("GoButton")
        self.horizontalLayout_4.addWidget(self.GoButton)
        self.horizontalLayout_3.addWidget(self.frame_16)
        self.verticalLayout.addWidget(self.frame_10)
        self.horizontalLayout.addWidget(self.frame_2)
        MainWindow.setCentralWidget(self.centralwidget)
        self.maintableURL.setVisible(False)
        self.delayURL.setVisible(False)
        self.delaytable.setVisible(False)

        self.FillButton.clicked.connect(self.Fill)
        self.GoButton.clicked.connect(self.Go)
        self.excelmaintableButton.clicked.connect(self.PB)
        self.exceldelaytableButton.clicked.connect(self.DB)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def TimeMatrix(self, jobsMatrix, seq):
        nb_machines = len(jobsMatrix)
        nb_jobs = len(seq)
        C = [[jobsMatrix[0][seq[0]]]]
        for i in range(1, nb_machines):
            C.append([C[-1][0] + jobsMatrix[i][seq[0]]])
        for j in range(1, nb_jobs):
            C[0].append(C[0][-1] + jobsMatrix[0][seq[j]])
        for j in range(1, nb_jobs):
            for i in range(1, nb_machines):
                C[i].append(max(C[i - 1][j], C[i][j - 1]) + jobsMatrix[i][seq[j]])
        return C

    def TimeMatrixwithprep(self, jobsMatrix, prepMatrix, seq):
            nb_machines = len(jobsMatrix)
            nb_jobs = len(seq)
            C = [[jobsMatrix[0][seq[0]] + prepMatrix[0][seq[0]][seq[0]]]]
            for i in range(1, nb_machines):
                    C.append([max(prepMatrix[i][seq[0]][seq[0]], C[-1][0]) + jobsMatrix[i][seq[0]]])
            for j in range(1, nb_jobs):
                    C[0].append(C[0][-1] + prepMatrix[0][seq[j - 1]][seq[j]] + jobsMatrix[0][seq[j]])
            for j in range(1, nb_jobs):
                    for i in range(1, nb_machines):
                            C[i].append(
                                    max(C[i - 1][j], C[i][j - 1] + prepMatrix[i][seq[j - 1]][seq[j]]) + jobsMatrix[i][
                                            seq[j]])
            return C

    def johnson(self,M1, M2):
        n_jobs = len(M1)
        jobsMatrix = [M1, M2]
        left_seq = [j for j in range(n_jobs) if jobsMatrix[0][j] <= jobsMatrix[1][j]]
        right_seq = [j for j in range(n_jobs) if jobsMatrix[0][j] > jobsMatrix[1][j]]
        left_seq = sorted(left_seq, key=lambda x: jobsMatrix[0][x])
        right_seq = sorted(right_seq, key=lambda x: jobsMatrix[1][x], reverse=True)
        optimalSeq = left_seq + right_seq
        return optimalSeq

    def cds(self,jobsMatrix):
            nb_machines = len(jobsMatrix)
            bestTime = float("inf")
            for k in range(1, nb_machines):
                    M1 = list(map(sum, zip(*jobsMatrix[:k])))
                    M2 = list(map(sum, zip(*jobsMatrix[-k:])))
                    seq = self.johnson(M1, M2)
                    timeMatrix = self.TimeMatrix(jobsMatrix, seq)
                    print(timeMatrix)
                    time = timeMatrix[-1][-1]
                    if time < bestTime:
                            bestSeq = seq
                            bestTime = time
            timeMatrix = self.TimeMatrix(jobsMatrix, bestSeq)
            print(timeMatrix, seq)
            return [bestSeq, bestTime, timeMatrix]

    def cds_prep(self,jobsMatrix,prepMatrix):
            nb_machines = len(jobsMatrix)
            bestTime = float("inf")
            for k in range(1, nb_machines):
                    M1 = list(map(sum, zip(*jobsMatrix[:k])))
                    M2 = list(map(sum, zip(*jobsMatrix[-k:])))
                    seq = self.johnson(M1, M2)
                    timeMatrix = self.TimeMatrixwithprep(jobsMatrix,prepMatrix, seq)
                    time = timeMatrix[-1][-1]
                    if time < bestTime:
                            bestSeq = seq
                            bestTime = time
            return [bestSeq, bestTime, timeMatrix]

    def cds_delay(self,jobsMatrix, delay):
            nb_machines = len(jobsMatrix)
            lowestTT = float("inf")
            for k in range(1, nb_machines):
                    M1 = list(map(sum, zip(*jobsMatrix[:k])))
                    M2 = list(map(sum, zip(*jobsMatrix[-k:])))
                    seq = self.johnson(M1, M2)
                    timeMatrix = self.TimeMatrix(jobsMatrix, seq)
                    TT = sum((x - y + abs(x - y)) / 2 for x, y in zip(timeMatrix[-1], delay))
                    if TT < lowestTT:
                            bestSeq = seq
                            lowestTT = TT
            time = timeMatrix[-1][-1]
            timeMatrix = self.TimeMatrix(jobsMatrix, seq)
            return [bestSeq, time, timeMatrix, lowestTT]

    def getdelay(self, endTime,delay):
        return [max(endTime[-1][i] - delay[i],0) for i in range(len(delay))]

    def getSequenceWithProperTime(self,jobsMatrix,prepMatrix):
        nb_machines=len(jobsMatrix)
        nb_jobs=len(jobsMatrix[0])
        seq=list(range(nb_jobs))
        TP=[0 for _ in range(nb_jobs)]
        for i in range(nb_jobs):
                TP[i]=[sum(list(zip(*jobsMatrix))[seq[i]])+sum(list(zip(*prepMatrix))[seq[i]][seq[1]]),i]
        TP.sort()
        return [x[1] for x in TP[::-1]]
        
    def TFR(self,table,seq,Cmax):
        return [sum([table[i][seq[j]] for j in range(len(table[0]))])/Cmax for i in range(len(table))]

    def TAP(self,prepMatrix,seq,Cmax):
        return [(prepMatrix[i][seq[0]][seq[0]]+sum(prepMatrix[i][seq[j-1]][seq[j]] for j in range(1,len(prepMatrix[0]))))/Cmax for i in range(len(prepMatrix))]

    def TAR(self,table,prepMatrix,seq,Cmax):
        tap=self.TAP(prepMatrix,seq,Cmax)
        tfr=self.TFR(table,seq,Cmax)
        return [1-x-y for x,y in zip(tfr,tap)]

    def preparationTable(self, n, m):

        self.prepMatrix=[[[]*n]*n]*m
        numOfTabs = self.tabWidget.count()
        while self.tabWidget.count()>1:
            lastTabIndex = self.tabWidget.count()
            self.tabWidget.removeTab(lastTabIndex-1)
        
        for j in range(m):
            
            table = QTableWidget()
            
            page = QWidget()
            page.layout = QVBoxLayout()
            page.layout.addWidget(table)
            page.setLayout(page.layout)
        
            self.tabWidget.addTab(page,"Machine "+str(j+1))
            
            tab = self.tabWidget.widget(j+1).children()[1]
            tab.setRowCount(n)
            tab.setColumnCount(n)
            tab.clear()
            
            headers=["Job "+str(i+1) for i in range(n)]
            
            tab.setHorizontalHeaderLabels(headers)
            tab.setVerticalHeaderLabels(headers)
            tab.horizontalHeader().setStretchLastSection(True)
            tab.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def preparationTables(self, n, m):
        #n, m = widget.nb_jobs, widget.nb_machines
        self.prepMatrix = []
    
        for machineTable in range(1,m+1):
            tab = self.tabWidget.widget(machineTable).children()[1]
            tabData = [[int(tab.item(row, col).text()) for col in range(n)] for row in range(n)]
            self.prepMatrix.append(tabData)

        return self.prepMatrix

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Flowshop Optimization Software"))
        self.maintableURL.setText(_translate("MainWindow", "From : "))
        item = self.maintable.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", "M1"))
        item = self.maintable.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "J1"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "maintable"))
        self.delayURL.setText(_translate("MainWindow", "From : "))
        item = self.delaytable.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", "Délai"))
        item = self.delaytable.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "J1"))
        self.Cmax.setText(_translate("MainWindow", "Cmax = "))
        self.prep.setText(_translate("MainWindow", "Avec Préparation"))
        self.delay.setText(_translate("MainWindow", "Avec Délai"))
        self.label.setText(_translate("MainWindow", "Nombre de Machines :"))
        self.label_2.setText(_translate("MainWindow", "Nombre de Jobs :"))
        self.excelmaintableButton.setText(_translate("MainWindow", "Parcourir Excel"))
        self.exceldelaytableButton.setText(_translate("MainWindow", "Parcourir Délai"))
        self.FillButton.setText(_translate("MainWindow", "Remplir"))
        self.comboBox.setItemText(0, _translate("MainWindow", "Gantt"))
        self.comboBox.setItemText(1, _translate("MainWindow", "TFR"))
        self.comboBox.setItemText(2, _translate("MainWindow", "TAR"))
        self.comboBox.setItemText(3, _translate("MainWindow", "TAP"))
        self.GoButton.setText(_translate("MainWindow", "Go"))

    def PB(self):
        self.maintableURL.setVisible(True)
        try:
            filter = 'Excel File (*.xlsx *.xls)'
            frame = QFileDialog.getOpenFileName(
                caption='Select a data file',
                directory=os.getcwd(),
                filter=filter,
                initialFilter='Excel File (*.xlsx *.xls)')
            self.maintableURL.setText(frame[0])

            wb = load_workbook(frame[0])
            ws = wb.active

            row = len(ws['A'])
            column = len(ws[1])

            table = [[0 for i in range(column)] for j in range(row)]

            for i in range(1, row + 1):
                for j in range(1, column + 1):
                    table[i - 1][j - 1] = ws[chr(64 + j) + str(i)].value

            self.maintable.setColumnCount(column)
            self.maintable.setRowCount(row)
            for i in range(row):
                item = QtWidgets.QTableWidgetItem()
                self.maintable.setVerticalHeaderItem(i, item)
            for j in range(column):
                item = QtWidgets.QTableWidgetItem()
                self.maintable.setHorizontalHeaderItem(j, item)
            _translate = QtCore.QCoreApplication.translate
            for i in range(row):
                item = self.maintable.verticalHeaderItem(i)
                item.setText(_translate("MainWindow", "Machine" + str(i + 1)))
            for j in range(column):
                item = self.maintable.horizontalHeaderItem(j)
                item.setText(_translate("MainWindow", "Job" + str(j + 1)))
            for i in range(row):
                for j in range(column):
                    self.maintable.setItem(i, j, QtWidgets.QTableWidgetItem(str(table[i][j])))



        except:
            self.maintableURL.setText("error")

    def DB(self):
        self.delay.setCheckState(True)
        self.delaytable.setVisible(True)
        self.delayURL.setVisible(True)
        try:
            filter = 'Block Note (*.txt);; Excel File (*.xlsx *.xls)'
            frame = QFileDialog.getOpenFileName(
                    caption='Select a data file',
                    directory=os.getcwd(),
                    filter=filter,
                    initialFilter='Excel File (*.xlsx *.xls)'
            )
            self.delayURL.setText(frame[0])

            wb = load_workbook(frame[0])
            ws = wb.active

            row = 1
            column = len(ws[1])

            table = [0 for i in range(column)]
            x = ""

            for i in range(1, column + 1):
                    table[i - 1] = ws[chr(64 + i) + str(1)].value
                    x += str(table[i - 1])+","
            self.delayURL.setText(x)

            print(table)

            row = len(ws['A'])
            column = len(ws[1])

            table = [[0 for i in range(column)] for j in range(row)]

            for i in range(1, row + 1):
                for j in range(1, column + 1):
                    table[i - 1][j - 1] = ws[chr(64 + j) + str(i)].value

            self.delaytable.setColumnCount(column)
            self.delaytable.setRowCount(row)
            for i in range(row):
                item = QtWidgets.QTableWidgetItem()
                self.delaytable.setVerticalHeaderItem(i, item)
            for j in range(column):
                item = QtWidgets.QTableWidgetItem()
                self.delaytable.setHorizontalHeaderItem(j, item)
            _translate = QtCore.QCoreApplication.translate
            for i in range(row):
                item = self.delaytable.verticalHeaderItem(i)
                item.setText(_translate("MainWindow", "Délai" + str(i + 1)))
            for j in range(column):
                item = self.delaytable.horizontalHeaderItem(j)
                item.setText(_translate("MainWindow", "J" + str(j + 1)))
            for i in range(row):
                for j in range(column):
                    self.delaytable.setItem(i, j, QtWidgets.QTableWidgetItem(str(table[i][j])))

        except:
            self.delayURL.setText("Data must be written in the First Row!")

    def Go(self):
        gantt = self.comboBox.currentText() == "Gantt"
        TFR = self.comboBox.currentText() == "TFR"
        TAR = self.comboBox.currentText() == "TAR"
        TAP = self.comboBox.currentText() == "TAP"

        C1 = not self.prep.isChecked() and not self.delay.isChecked()
        C2 = self.prep.isChecked() and not self.delay.isChecked()
        C3 = not self.prep.isChecked() and self.delay.isChecked()
        C4 = self.prep.isChecked() and self.delay.isChecked()

        row = self.maintable.rowCount()
        column = self.maintable.columnCount()

        color = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']
        prepMatrix = [[[1 for i in range(column)] for j in range(column)] for k in range(row)]
        delayarr = list(map(int, self.delayURL.text().split(",")[:-1]))
        table = [[0 for i in range(column)] for j in range(row)]
        try:
            for i in range(row):
                for j in range(column):
                    table[i][j] = int(self.maintable.item(i, j).text())
        except:
            table = [[0 for i in range(column)] for j in range(row)]

        print('cringe')
        print(table)
        print('parabolus')

        if C1:
            print('nPnD')

            res = self.cds(table)
            seq= res[0]
            endTime=res[-1]
            self.Cmax.setText("Cmax = " + str(endTime[-1][-1]))

        if C2:
            print('PnD')

            prepMatrix = self.preparationTables(column, row)
            print(prepMatrix)

            seq = self.getSequenceWithProperTime(table,prepMatrix)
            res = self.TimeMatrixwithprep(table,prepMatrix,seq)
            print(res)
            endTime=res
            print(endTime)
            self.Cmax.setText("Cmax = " + str(endTime[-1][-1]))

        if C3:
            print('nPD')

            res = self.cds_delay(table,delayarr)
            print(res)
            seq = res[0]
            endTime = res[-2]
            print(endTime)
            self.Cmax.setText("Cmax = " + str(endTime[-1][-1]))

        if C4:
            print('PD')

            prepMatrix = self.preparationTables(column, row)
            print(prepMatrix)

            seq = self.getSequenceWithProperTime(table,prepMatrix)
            endTime = self.TimeMatrixwithprep(table,prepMatrix,seq)
            delay = self.getdelay(endTime,delayarr)
            print(endTime)
            print(delay)
            self.Pexcelfile.setText('Total Tardiness : '+str(sum(delay)))
            self.Cmax.setText("Cmax = " + str(endTime[-1][-1]))

        if gantt:
            machine = self.maintable.rowCount()
            job = self.maintable.columnCount()
            table = [[0 for i in range(job)] for j in range(machine)]

            for i in range(machine):
                for j in range(job):
                    try:
                        table[i][j] = int(self.maintable.item(i, j).text())
                    except:
                        continue

            fig, gnt = plt.subplots()
            gnt.set_ylim(0, len(table) * 10)
            gnt.set_xlim(0, endTime[-1][-1])

            gnt.set_yticks([10*(i+1) - 5 for i in range(len(table))])
            gnt.set_yticklabels(["M"+str(len(table)-i) for i in range(len(table))])

            for i in range(len(table[0])):
                for j in range(len(table)):
                    gnt.broken_barh([(endTime[j][i] - table[j][seq[i]], table[j][seq[i]])],(10 *(len(table) - j - 1) + 1, 5),facecolors=('tab:' + str(color[i])))
            
            p_ja = mpatches.Patch(color='blue', label='Job1')
            p_jb = mpatches.Patch(color='orange', label='Job2')
            p_jc = mpatches.Patch(color='green', label='Job3')
            p_jd = mpatches.Patch(color='red', label='Job4')
            p_je = mpatches.Patch(color='purple', label='Job5')
            p_jf = mpatches.Patch(color='brown', label='Job6')
            p_jg = mpatches.Patch(color='pink', label='Job7')
            p_jh = mpatches.Patch(color='gray', label='Job8')
            p_ji = mpatches.Patch(color='olive', label='Job9')
            p_jj = mpatches.Patch(color='cyan', label='Job10')

            x = [p_ja, p_jb, p_jc, p_jd, p_je, p_jf, p_jg, p_jh, p_ji, p_jj]
            f = [x[i] for i in range(column)]

            plt.legend(handles=f, loc='best')

            plt.show()

        if TFR:
                x = self.TFR(table,seq,endTime[-1][-1])
                machines = list(range(1,len(table)+1))
                plt.bar(machines,x,align='center')
                plt.xlabel(["M"+str(i+1) for i in range(len(table))])
                for i in range(len(x)):
                        plt.hlines(x[i],0,machines[i])
                plt.show()

        if TAR:
                x = self.TAR(table,prepMatrix,seq,endTime[-1][-1])
                print(x)
                machines = list(range(1,len(table)+1))
                plt.bar(machines,x,align='center')
                plt.xlabel(["M"+str(i+1) for i in range(len(table))])
                for i in range(len(x)):
                        plt.hlines(x[i],0,machines[i])
                plt.show()

        if TAP and (C2 or C4):
                print(prepMatrix)
                x = self.TAP(prepMatrix,seq,endTime[-1][-1])
                print(x)
                machines = list(range(1,len(table)+1))
                plt.bar(machines,x,align='center')
                plt.xlabel(["M"+str(i+1) for i in range(len(table))])
                for i in range(len(x)):
                        plt.hlines(x[i],0,machines[i])
                plt.show()

    def Fill(self):
        row = self.spinBox.value()
        column = self.spinBox_2.value()

        self.maintable.setColumnCount(column)
        self.maintable.setRowCount(row)

        for i in range(row):
            item = QtWidgets.QTableWidgetItem()
            self.maintable.setVerticalHeaderItem(i, item)
        for j in range(column):
            item = QtWidgets.QTableWidgetItem()
            self.maintable.setHorizontalHeaderItem(j, item)
        _translate = QtCore.QCoreApplication.translate
        for i in range(row):
            item = self.maintable.verticalHeaderItem(i)
            item.setText(_translate("MainWindow", "Machine"+str(i + 1)))
        for j in range(column):
            item = self.maintable.horizontalHeaderItem(j)
            item.setText(_translate("MainWindow", "Job"+str(j+1)))

        if self.prep.isChecked():
            self.preparationTable(column, row)

        if self.delay.isChecked():
            self.delaytable.setVisible(True)

            self.delaytable.setColumnCount(column)
            self.delaytable.setRowCount(1)

            for j in range(column):
                item = QtWidgets.QTableWidgetItem()
                self.delaytable.setHorizontalHeaderItem(j, item)
            _translate = QtCore.QCoreApplication.translate
            for j in range(column):
                item = self.delaytable.horizontalHeaderItem(j)
                item.setText(_translate("MainWindow", "Job"+str(j+1)))

        elif not self.delay.isChecked():
            self.delayURL.setVisible(False)
            self.delaytable.setVisible(False)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
